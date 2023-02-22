import os
import re
import sys
from utils import *
import time
from openpyxl import Workbook


class LogExtractor:

    def __init__(self):
        # for bebugging purpose
        # self.testing = []

        # static machine works between these two events only star and stop
        self.is_nr10 = False
        self.is_nr12 = True

        # data that will be collected during the execution and later paste the data to output files
        self.final_data = []
        self.text_data = []
        self.date_matcher = re.compile(r'[2][0-9]{3}(.+)0x[a-zA-Z0-9]{4}|\s+[2][0-9]{3}(.+)0x[a-zA-Z0-9]{4}')
        self.nr10_pattern_not_acceptable = re.compile(r'\s+Serving\s+Cell\s+PCI\s+=\s+NA\n|'
                                                      r'Serving\s+Cell\s+PCI\s+=\s+NA\n')
        self.nr11_pattern_one = re.compile(r'Serving\s+Cell\s+PCI\s+=\s+0\n|\s+Serving\s+Cell\s+PCI\s+=\s+0\n')
        self.nr11_pattern_two = re.compile(r'Serving\s+Cell\s+PCI\s+=\s+1\n|\s+Serving\s+Cell\s+PCI\s+=\s+1\n')
        self.lt21_pattern_one = re.compile(r'\s+Physical\s+Cell\s+ID\s+=\s+0\n|Physical\s+Cell\s+ID\s+=\s+0\n')
        self.lt21_pattern_two = re.compile(r'\s+Physical Cell ID\s+=\s+1\n|Physical\s+Cell\s+ID\s+=\s+1\n')

    def create_table(self, data: dict, file_path: str) -> None:
        # Create a new workbook
        workbook = Workbook()
        # Get the active sheet
        sheet = workbook.active
        # Adding headers
        headers = list(data.keys())
        for i, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=i, value=header)

        # Adding data
        for i, lst in enumerate(data.values(), start=1):
            for j, value in enumerate(lst, start=2):
                sheet.cell(row=j, column=i, value=value)
        # Save the workbook
        workbook.save(file_path)
        # create table
        # tab = sheet.add_table(
        #     ref='A1:{}'.format(openpyxl.utils.get_column_letter(len(headers)) + str(len(data.values()) + 1)),
        #     name='mytable')
        # tab.style = 'Table Style Light 11'

    def get_output_data(self, output_file_path: str) -> dict:
        with open(output_file_path) as file:
            lines = file.readlines()
            data = {}
            for line in lines:
                match = re.match(self.date_matcher, line)
                if match:
                    if data.get("time", None):
                        data["time"].append(line[:25])
                    else:
                        data["time"] = [line[:25]]

                else:
                    if "=" in line:
                        key, value = line.split("=")
                        key, value = key.strip(), value.strip()
                        if data.get(key):
                            if len(data[key]) != len(data['time']) - 1:
                                while len(data[key]) < len(data['time']) - 1:
                                    data[key].append(None)
                            data[key].append(value)
                        else:
                            data[key] = []
                            if len(data[key]) != len(data['time']) - 1:
                                while len(data[key]) < len(data['time']) - 1:
                                    data[key].append(None)
                            data[key].append(value)
            return data

    def get_headers(self, chunk_data: str) -> list | None:
        pattern = re.compile(r"\s*--+\n(.+)\n\s*--+\n", re.DOTALL)
        header_list = []
        for data in pattern.finditer(chunk_data):
            rows_data = data.group().split("\n")[:-1]
            while rows_data:
                rows = []
                counter = 0
                while counter < 2:
                    rows.append(rows_data[0].strip())
                    if "---" in rows_data[0].strip():
                        counter += 1
                    if counter == 0:
                        rows.clear()
                    del rows_data[0]
                original_list = [row.split("|")[1:-1] for row in rows[1:-1]]
                max_count = max([len(item) for item in original_list])
                original_list = [item for item in original_list if len(item) == max_count]
                headers = [" ".join(map(str.strip, item)) for item in zip(*original_list)]
                header_list.append(headers)
                rows.clear()
            return header_list

    def get_rows(self, chunk_data: str) -> list | None:
        pattern = re.compile(r"\s*--+\n(.+)\n\s*--+\n(.+)\|\n\n", re.DOTALL)
        data = pattern.search(chunk_data)
        result = []
        counter = 0
        if data:
            rows_data = data.group().split("\n")[:-1]
            while rows_data:
                if rows_data[0].strip().startswith("----"):
                    counter += 1
                if counter != 2:
                    rows_data.pop(0)
                if counter == 2:
                    if rows_data[0].strip().endswith("|") and rows_data[0].strip().startswith("|"):
                        result.append([item.strip() for item in rows_data[0].strip().split("|") if item])
                        counter = 0
                    else:
                        rows_data.pop(0)
        return result

    def get_file_name(self, file_name: str, output_path: str) -> str:
        """
        Function will return the absolute location of the output file.
        """
        file_abs_path: str = os.path.join(output_path, file_name)
        destination_dir: str = os.path.dirname(file_abs_path)
        if not os.path.exists(destination_dir):
            os.makedirs(destination_dir)
        return file_abs_path

    def read_input_file(self, input_path: str, output_path: str) -> None:
        """
        Function will execute the script and create output files
        :param input_path: Absolute location of the directory of input files that needs to be read
        :param output_path: Absolute location of the directory of output files that needs to be updated
        :return: None
        """
        if os.path.exists(input_path):
            for file_name in os.listdir(input_path):
                start_time = time.time()
                # reading each text file from the directory
                if file_name.lower().endswith(".txt"):
                    # joining the input directory with the text file to get absolute location
                    file_abs_path: str = os.path.join(input_path, file_name)
                    with open(file_abs_path) as f:
                        self.text_data: list = f.readlines()

                        # Generating output text file name using input text file name
                        output_file_name: str = file_name[:len(file_name) - 4] + "_output.txt"

                        self.final_data.append(f"Source file: {file_name}\n")
                        self.final_data.append(f"Output File: {output_file_name}\n")

                        # QCAT version needs to be added as specified in the text file once it
                        # is done, it should break the loop and go to the next section
                        for i in range(10):
                            line = self.text_data[i]
                            if "%QCAT VERSION" in line:
                                self.final_data.append(line)
                                break

                        # blocking an index of 3 for capturing execution time for each file.
                        self.final_data.append("TimeTaken\n")

                        res = []
                        end_idx = start_idx = counter = b9be_time = b97f_time = nr11_count = 0
                        lt21_res = []
                        for idx, line in enumerate(self.text_data):
                            match = re.match(self.date_matcher, line)
                            if match and counter == 0:
                                start_idx = idx
                                counter = 1
                            elif match and counter == 1:
                                end_idx = idx
                                res = self.text_data[start_idx:end_idx]
                                start_idx = idx

                                # Trying to capture lt21 using Cell ID from 0 to 1 or 1 to 0
                                # condition should be matched within nr10 and nr12
                                if "0xB193" in res[0]:
                                    if (any(re.match(self.lt21_pattern_one, item) for item in res)
                                        or any(re.match(self.lt21_pattern_two, item) for item in res)) \
                                            and len(lt21_res) == 0:
                                        lt_21_data = [item for item in res if "Physical Cell ID" in item]
                                        _, value = lt_21_data[0].split("=")
                                        _, value = _, value.strip()
                                        lt21_res.append(value)

                                    # if Cell ID is changing and validator has correct Cell ID {0, 1}
                                    # from previous capture
                                    elif ((any(re.match(self.lt21_pattern_two, item) for item in res)
                                           and lt21_res[0] == '0')
                                          or (any(re.match(self.lt21_pattern_one, item) for item in res)
                                              and lt21_res[0] == '1')) and self.is_nr10 and not self.is_nr12:
                                        time_stamp = res[0][:39]
                                        self.final_data.append(time_stamp + "\n")
                                        self.final_data.append(f"Metric = LT21" + "\n")
                                        for line_txt in res:
                                            if "Physical Cell ID" in line_txt:
                                                self.final_data.append(line_txt.strip() + "\n")
                                            elif "ARFCN" in line_txt:
                                                self.final_data.append(line_txt.strip() + "\n")
                                        lt21_res.clear()
                                    # if Cell ID is continuously repeating for either 0 or 1, then it just skip
                                    # that chunk
                                    elif ((any(re.match(self.lt21_pattern_two, item) for item in res)
                                           and lt21_res[0] == '1')
                                          or (any(re.match(self.lt21_pattern_one, item) for item in res)
                                              and lt21_res[0] == '0')):
                                        continue
                                    # if Cell ID anything apart from 0 or 1, it should clear the validator
                                    else:
                                        lt21_res.clear()

                                # Capturing nr10 to start the machine
                                elif self.is_nr12 and not self.is_nr10:
                                    if "0xB9BE" in res[0]:
                                        b9be_time = time_diff(res[0])
                                    elif "0xB97F" in res[0] and time_diff(res[0]) - b9be_time <= 1.0 \
                                            and any("Serving Cell PCI" in item for item in res) \
                                            and not any(re.match(self.nr10_pattern_not_acceptable, item)
                                                        for item in res):
                                        pci_match = [item for item in res if "Serving Cell PCI" in item][0]
                                        key, value = pci_match.split("=")
                                        key, value = key.strip(), value.strip()
                                        if value.isnumeric():
                                            b97f_time = time_diff(res[0])
                                            time_stamp = res[0][:39]
                                            self.final_data.append(time_stamp + "\n")
                                            self.final_data.append(f"Metric = NR10" + "\n")
                                            for line_txt in res:
                                                if "Raster ARFCN" in line_txt:
                                                    self.final_data.append(line_txt.strip() + "\n")
                                                elif "Serving Cell PCI" in line_txt:
                                                    self.final_data.append(line_txt.strip() + "\n")
                                    elif "0xB825" in res[0] and time_diff(res[0]) - b97f_time <= 1.0:
                                        cols = self.get_headers("".join(res))
                                        rows = self.get_rows("".join(res))
                                        if rows and cols:
                                            final_row = [item for row in rows for item in row]
                                            final_column = [item for col in cols for item in col]
                                            dict_data = dict(zip(final_column, final_row))
                                            self.final_data.append(f"Band Number = {dict_data['Band Number']}" + "\n")
                                            b9be_time = 0
                                            b97f_time = 0
                                            self.is_nr12 = False
                                            self.is_nr10 = True
                                    elif "0xB97F" in res[0] \
                                            and "Serving Cell PCI" in self.final_data[len(self.final_data) - 1]:
                                        continue
                                    elif time_diff(res[0]) - b97f_time > 1.0 and b97f_time != 0:
                                        nr10_counter = 4
                                        while nr10_counter > 0:
                                            self.final_data.pop()
                                            nr10_counter -= 1

                                # Capturing nr11 and nr12
                                if not self.is_nr12 and self.is_nr10:
                                    if "0xB9BE" in res[0]:
                                        b9be_time = time_diff(res[0])
                                    elif "0xB97F" in res[0] and any("Serving Cell PCI" in item for item in res) \
                                            and any(re.match(self.nr11_pattern_one, item) for item in res) \
                                            and nr11_count == 0:
                                        nr11_count = 1
                                    elif not (any(re.match(self.nr11_pattern_one, item) for item in res)
                                              or any(re.match(self.nr11_pattern_two, item) for item in res)) \
                                            and "0xB97F" in res[0] and nr11_count == 1:
                                        nr11_count = 0
                                    elif "0xB97F" in res[0] and time_diff(res[0]) - b9be_time <= 1.0 \
                                            and any("Serving Cell PCI" in item for item in res) \
                                            and nr11_count == 1 \
                                            and any(re.match(self.nr11_pattern_two, item) for item in res):
                                        b97f_time = time_diff(res[0])
                                        time_stamp = res[0][:39]
                                        self.final_data.append(time_stamp + "\n")
                                        self.final_data.append(f"Metric = NR11" + "\n")
                                        nr11_count = 0
                                        for line_txt in res:
                                            # self.testing.append(line_txt)
                                            if "Raster ARFCN" in line_txt:
                                                self.final_data.append(line_txt.strip() + "\n")
                                            elif "Serving Cell PCI" in line_txt:
                                                self.final_data.append(line_txt.strip() + "\n")
                                    elif "0xB97F" in res[0] and any(
                                            re.match(self.nr11_pattern_one, item) for item in res) \
                                            and nr11_count == 1:
                                        continue
                                    elif "0xB825" in res[0] and time_diff(res[0]) - b97f_time <= 1.0:
                                        # print(self.text_data.index(res[0]))
                                        cols = self.get_headers("".join(res))
                                        rows = self.get_rows("".join(res))
                                        if rows and cols:
                                            final_row = [item for row in rows for item in row]
                                            final_column = [item for col in cols for item in col]
                                            dict_data = dict(zip(final_column, final_row))
                                            self.final_data.append(f"Band Number = {dict_data['Band Number']}" + "\n")
                                            b9be_time = 0
                                            b97f_time = 0
                                    elif time_diff(res[0]) - b97f_time > 1.0 \
                                            and "Serving Cell PCI" in self.final_data[len(self.final_data) - 1]:
                                        validtor = 4
                                        while validtor > 0:
                                            self.final_data.pop()
                                            validtor -= 1

                                    elif "0x1FFB" in res[0] \
                                            and (any("RRC State = Closing" in item for item in res)
                                                 or any("ID=3326" in item for item in res)) \
                                            and not self.is_nr12:
                                        time_stamp = res[0][:39]
                                        self.final_data.append(time_stamp + "\n")
                                        self.final_data.append(f"Metric = NR12" + "\n")
                                        self.is_nr12 = True
                                        self.is_nr10 = False

                    end_time = time.time()
                    self.final_data[3] = f"Execution took approx: {round((end_time - start_time), 3)} seconds \n"

                    with open(self.get_file_name(output_file_name, output_path), "w") as file:
                        for line in self.final_data:
                            file.write(line)

                    self.final_data.clear()

                    # creating Excel file corresponding to the input text file name
                    output_excel_file_name: str = file_name[:len(file_name) - 4] + "_output.xlsx"

                    # Extracting all data from output text file in a dictionary
                    data = self.get_output_data(self.get_file_name(output_file_name, output_path))

                    # saving all dictionary data to the excel file.
                    self.create_table(data, self.get_file_name(output_excel_file_name, output_path))


if __name__ == '__main__':
    start_time_log = time.time()
    input_txt_files_dir = sys.argv[1]
    output_txt_files_dir = sys.argv[2]
    log_extractor = LogExtractor()
    log_extractor.read_input_file(input_txt_files_dir, output_txt_files_dir)
    end_time_log = time.time()
    print(f"Execution took approximately {round(end_time_log - start_time_log, 3)} seconds")
